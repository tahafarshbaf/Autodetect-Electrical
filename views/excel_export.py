"""
Excel export module for the YOLO detection app.

Fills detection results into a company BOQ (Bill of Quantities) template.
The template has a repeating block structure (one block per panel),
each block holding up to 30 element rows.

Import and call fill_template() from app.py.
"""

import re
from io import BytesIO
import openpyxl

from amp_ocr import AMPERAGE_KEY_SEPARATOR

# ---------------------------------------------------------------------------
# Template layout constants
# Adjust these if the template structure ever changes.
# ---------------------------------------------------------------------------
BLOCK_HEIGHT = 41          # rows between the start of one block and the next
DATA_ROWS_PER_BLOCK = 30   # element rows available per block
HEADER_OFFSET = 4          # column header row is 4 rows after block start
DATA_START_OFFSET = 5      # first data row is 5 rows after block start
DRAW_NO_ROW_OFFSET = 2     # DRAW NO row is 2 rows after each block's start row

COL_DESCRIPTION = 2   # column B — element name
COL_RANGE = 3         # column C — specification (e.g. 3-Pole / 1-Pole)
COL_QTY = 8            # column H — quantity


# ---------------------------------------------------------------------------
# Priority order for the Excel output.
#
# Element names (the part before the first digit, e.g. "MCB" from "MCB1P")
# listed here appear FIRST in the Excel file, in this exact order —
# regardless of the order they were detected in the images.
#
# Any element name NOT in this list is placed after all listed ones,
# sorted alphabetically among themselves.
#
# Edit this list to match your own priority order. Matching is
# case-insensitive.
# ---------------------------------------------------------------------------
CLASS_PRIORITY_ORDER = [
    "MCCB",
    "MCB",
    "Contactor",
    "RCD",
    "Relay",
    "Fuse",
]


# ---------------------------------------------------------------------------
# Fixed trailing rows.
#
# These are always appended as the LAST rows of the export, in this exact
# order, regardless of what was detected/calculated elsewhere (Detection
# page, Terminal page). Their quantity is left blank so it can be filled
# in by hand afterwards. They're independent from the "Busbar<size>mm²"
# rows the Terminal page may add (those come from the wire/terminal size
# calculation and aren't related to this fixed "Busbar" line item).
#
# Edit this list if the set/order of fixed rows ever needs to change.
# ---------------------------------------------------------------------------
FIXED_TRAILING_ITEMS = ["Wire", "Cubicle", "Busbar", "Installation"]


# Matches everything before the first digit as the name, and the first
# digit onward as the spec, e.g. "MCB1P" -> ("MCB", "1P").
_CLASS_NAME_PATTERN = re.compile(r"^([^\d]+)(\d.*)$")

# Matches an amperage suffix appended by the Detection page's OCR step,
# e.g. "MCB3P__63A" -> ("MCB3P", "63A"). See amp_ocr.build_class_key().
_AMPERAGE_SUFFIX_PATTERN = re.compile(
    rf"^(.*){re.escape(AMPERAGE_KEY_SEPARATOR)}(\d+(?:\.\d+)?A)$"
)


def split_class_name(class_name: str):
    """
    Splits a YOLO class name into (element_name, spec), where everything
    before the first digit is the element name and the first digit
    onward is the spec. Also decodes an optional amperage suffix added
    by the Detection page's OCR step (see amp_ocr.build_class_key()),
    appending the amperage to the spec so it ends up in the same "Range"
    column of the BOQ template.

    Examples:
        "MCB1P"       -> ("MCB", "1P")
        "Contactor3P" -> ("Contactor", "3P")
        "Relay"       -> ("Relay", "")            # no digit found
        "MCB3P__63A"  -> ("MCB", "3P 63A")         # amperage from OCR
        "Relay__12A"  -> ("Relay", "12A")
    """
    amperage_suffix = None
    amperage_match = _AMPERAGE_SUFFIX_PATTERN.match(class_name)
    if amperage_match:
        class_name, amperage_suffix = amperage_match.group(1), amperage_match.group(2)

    match = _CLASS_NAME_PATTERN.match(class_name)
    if match:
        name, spec = match.group(1).strip(), match.group(2).strip()
    else:
        name, spec = class_name, ""

    if amperage_suffix:
        spec = f"{spec} {amperage_suffix}".strip()

    return name, spec


def _priority_sort_key(class_name: str):
    """
    Sort key that places element names according to CLASS_PRIORITY_ORDER
    first (in that exact order), and any unlisted element names after,
    sorted alphabetically. Ties within the same element name are broken
    alphabetically by the full class name (so e.g. MCB1P sorts before MCB3P).
    """
    element_name, _ = split_class_name(class_name)
    element_name_upper = element_name.upper()

    priority_lookup = {name.upper(): i for i, name in enumerate(CLASS_PRIORITY_ORDER)}

    if element_name_upper in priority_lookup:
        return (0, priority_lookup[element_name_upper], class_name)
    return (1, element_name_upper, class_name)


def _build_merge_anchor_map(ws):
    """
    Precomputes a {coordinate: anchor_cell} map for every cell that falls
    inside a merged range, so repeated lookups (e.g. once per item row)
    don't have to linear-scan ws.merged_cells.ranges every time.
    """
    anchor_map = {}
    for merged_range in ws.merged_cells.ranges:
        anchor = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                anchor_map[(row, col)] = anchor
    return anchor_map


def _writable_cell(ws, row: int, column: int, merge_anchor_map: dict):
    """
    Returns the cell that should actually be written to for (row, column).

    If that position falls inside a merged range, openpyxl only allows
    writing to that range's top-left ("anchor") cell — every other cell
    in the range is a read-only MergedCell. merge_anchor_map (built once
    via _build_merge_anchor_map) redirects to the anchor automatically so
    callers never have to worry about hitting a MergedCell directly.
    """
    if (row, column) in merge_anchor_map:
        return merge_anchor_map[(row, column)]
    return ws.cell(row=row, column=column)


def fill_template(
    template_file,
    class_totals: dict,
    panel_name: str = "",
    date: str = "",
    client_name: str = "",
    draw_no: str = "",
    start_page: int = 1,
):
    """
    Fills the BOQ template with detection results.

    Args:
        template_file: path to the .xlsx template, or a file-like object
                        (e.g. a Streamlit UploadedFile).
        class_totals: dict of {class_name: total_count}, e.g.
                       {"CB_3Pole": 12, "CB_1Pole": 8, ...}
        panel_name: optional text to write into the "Panel Name" field.
        date: optional text to write into the "Date" field.
        client_name: optional text to write into the "To: Client" field
                     (the "به: شركت" field in the template).
        draw_no: optional text (e.g. "DRAW NO: 1405-92") to write into
                 the DRAW NO row of the block for start_page (i.e. the
                 same page panel_name/date/client_name are written to —
                 NOT a fixed absolute cell).
        start_page: which page (1-indexed) to start writing into. Page 1 is
                    the first block (rows 1-41), page 2 is the second
                    (rows 42-82), and so on. If there are more element
                    types than fit on one page, writing continues onto
                    the following page(s).

    The rows in FIXED_TRAILING_ITEMS (Wire, Cubicle, Busbar, Installation)
    are always appended as the last rows of the export, in that exact
    order, with a blank quantity for manual entry afterwards — regardless
    of what class_totals contains.

    Returns:
        A BytesIO object containing the filled .xlsx file, ready for download.

    Raises:
        ValueError: if start_page is out of range, or if there are more
                    distinct element types than the remaining pages in the
                    template can hold.
    """
    wb = openpyxl.load_workbook(template_file)
    ws = wb.active
    merge_anchor_map = _build_merge_anchor_map(ws)

    # Sort by the priority list defined above (CLASS_PRIORITY_ORDER), not by
    # detection order or count.
    items = sorted(class_totals.items(), key=lambda x: _priority_sort_key(x[0]))

    # Always append the fixed trailing rows (blank quantity) at the very
    # end, after everything else, in FIXED_TRAILING_ITEMS's exact order.
    items = items + [(name, None) for name in FIXED_TRAILING_ITEMS]

    total_pages_available = count_available_blocks(ws)

    if start_page < 1 or start_page > total_pages_available:
        raise ValueError(
            f"Invalid page number: {start_page}. This template has "
            f"{total_pages_available} page(s), so start_page must be "
            f"between 1 and {total_pages_available}."
        )

    pages_needed = max(1, -(-len(items) // DATA_ROWS_PER_BLOCK))  # ceil division
    last_page_used = start_page + pages_needed - 1

    if last_page_used > total_pages_available:
        raise ValueError(
            f"Not enough pages left in the template: {len(items)} element types "
            f"(including the {len(FIXED_TRAILING_ITEMS)} fixed trailing rows) "
            f"need {pages_needed} page(s) of {DATA_ROWS_PER_BLOCK} rows each, "
            f"starting from page {start_page} that would require pages up to "
            f"{last_page_used}, but the template only has {total_pages_available} "
            f"page(s) in total."
        )

    # DRAW NO is written once per file, into the DRAW NO row of the block
    # for start_page — i.e. the same page the rest of the header info
    # (panel_name/date/client_name) is written to.
    if draw_no:
        draw_no_row = block_start_row(start_page) + DRAW_NO_ROW_OFFSET
        _writable_cell(ws, draw_no_row, 2, merge_anchor_map).value = draw_no

    item_index = 0
    for page_offset in range(pages_needed):
        page_number = start_page + page_offset
        page_start_row = block_start_row(page_number)
        data_start_row = page_start_row + DATA_START_OFFSET

        if panel_name:
            _writable_cell(ws, page_start_row + 1, 1, merge_anchor_map).value = f"Panel Name: {panel_name}"
        if date:
            _writable_cell(ws, page_start_row, 1, merge_anchor_map).value = f"Date : {date}"
        if client_name:
            _writable_cell(ws, page_start_row + 1, 8, merge_anchor_map).value = f"به : {client_name}"

        for row_offset in range(DATA_ROWS_PER_BLOCK):
            if item_index >= len(items):
                break

            class_name, count = items[item_index]
            element_name, spec = split_class_name(class_name)

            row = data_start_row + row_offset
            _writable_cell(ws, row, COL_DESCRIPTION, merge_anchor_map).value = element_name
            _writable_cell(ws, row, COL_RANGE, merge_anchor_map).value = spec
            # count is None for the fixed trailing rows (Wire, Cubicle,
            # Busbar, Installation) so their quantity cell is left blank
            # for manual entry.
            _writable_cell(ws, row, COL_QTY, merge_anchor_map).value = count

            item_index += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def count_available_blocks(ws):
    """Counts how many repeating blocks (pages) exist in the template by
    scanning for 'ITEM' header cells in column A."""
    count = 0
    row = 5
    while row <= ws.max_row:
        if ws.cell(row=row, column=1).value == "ITEM":
            count += 1
        row += BLOCK_HEIGHT
    return count


def count_pages_in_template(template_file):
    """
    Convenience function: opens the template and returns how many pages
    (panel blocks) it contains, without modifying anything.
    Useful for showing the user "Page 1 to N" in the UI.
    """
    wb = openpyxl.load_workbook(template_file, read_only=True)
    ws = wb.active
    return count_available_blocks(ws)


def block_start_row(page_number: int) -> int:
    """
    Returns the row number where the given page (1-indexed) starts.
    E.g. page 1 starts at row 1, page 2 at row 42, page 3 at row 83, etc.
    """
    return 1 + (page_number - 1) * BLOCK_HEIGHT