"""
Reads the "panel price" Excel file: a small, user-uploaded workbook that
holds just two values used to fill the Word proposal (TFP) template —

    PRICE_CELL        -> the panel price (a number)
    PANEL_COUNT_CELL  -> the number of panels

This file is uploaded fresh from the UI every time (like the BOQ Excel
template), so this module doesn't need a fixed path.

PLACEHOLDER cell references below — update PRICE_CELL / PANEL_COUNT_CELL
(and SHEET_NAME if it's not the active sheet) to the real cell addresses
once you've confirmed them in the actual file. Everything else
(tfp_generator.py, views/export.py) reads through this module, so you
only need to change it in one place.
"""

import openpyxl

# None = use the workbook's active sheet. Set to a string (e.g. "Sheet1")
# if the values live on a specific named sheet instead.
SHEET_NAME = None

# TODO: set these to the real cell addresses in the panel-price file.
PRICE_CELL = "F7"
PANEL_COUNT_CELL = "D7"


def read_panel_price(file_path_or_obj):
    """
    Opens the panel-price Excel file and returns (price, panel_count)
    read from PRICE_CELL / PANEL_COUNT_CELL.

    Args:
        file_path_or_obj: path to the .xlsx file, or a file-like object
                           (e.g. a Streamlit UploadedFile).

    Returns:
        (price, panel_count) as read from the workbook (whatever type
        openpyxl gives back — usually int/float).

    Raises:
        ValueError: if either cell is empty, so the caller can show a
                    clear message instead of silently generating a
                    document with missing values.
    """
    # data_only=True: read the last-calculated VALUE of each cell rather
    # than its formula, in case PRICE_CELL/PANEL_COUNT_CELL hold formulas.
    wb = openpyxl.load_workbook(file_path_or_obj, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME else wb.active

    price = ws[PRICE_CELL].value
    panel_count = ws[PANEL_COUNT_CELL].value

    if price in (None, ""):
        raise ValueError(
            f"Cell {PRICE_CELL} (price) is empty in the panel-price file. "
            f"Check PRICE_CELL in panel_price.py."
        )
    if panel_count in (None, ""):
        raise ValueError(
            f"Cell {PANEL_COUNT_CELL} (panel count) is empty in the "
            f"panel-price file. Check PANEL_COUNT_CELL in panel_price.py."
        )

    return price, panel_count