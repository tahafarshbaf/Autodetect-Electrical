"""
Excel export module for the YOLO detection app.

Fills detection results into a company BOQ (Bill of Quantities) template.
The template has a repeating block structure (one block per panel),
each block holding up to 30 element rows.

Import and call fill_template() from app.py.
"""

import re
from io import BytesIO
import openpyxl

# ---------------------------------------------------------------------------
# Template layout constants
# Adjust these if the template structure ever changes.
# ---------------------------------------------------------------------------
BLOCK_HEIGHT = 41          # rows between the start of one block and the next
DATA_ROWS_PER_BLOCK = 30   # element rows available per block
HEADER_OFFSET = 4          # column header row is 4 rows after block start
DATA_START_OFFSET = 5      # first data row is 5 rows after block start

COL_DESCRIPTION = 2   # column B — element name
COL_RANGE = 3         # column C — specification (e.g. 3-Pole / 1-Pole)
COL_QTY = 8            # column H — quantity


# ---------------------------------------------------------------------------
# Priority order for the Excel output.
#
# Element names (the part before the first digit, e.g. "MCB" from "MCB1P")
# listed here appear FIRST in the Excel file, in this exact order —
# regardless of the order they were detected in the images.
#
# Any element name NOT in this list is placed after all listed ones,
# sorted alphabetically among themselves.
#
# Edit this list to match your own priority order. Matching is
# case-insensitive.
# ---------------------------------------------------------------------------
CLASS_PRIORITY_ORDER = [
    "MCCB",
    "MCB",
    "Contactor",
    "RCD",
    "Relay",
    "Fuse",
]


# Matches everything before the first digit as the name, and the first
# digit onward as the spec, e.g. "MCB1P" -> ("MCB", "1P").
_CLASS_NAME_PATTERN = re.compile(r"^([^\d]+)(\d.*)$")


def split_class_name(class_name: str):
    """
    Splits a YOLO class name into (element_name, spec), where everything
    before the first digit is the element name and the first digit
    onward is the spec.

    Examples:
        "MCB1P"       -> ("MCB", "1P")
        "Contactor3P" -> ("Contactor", "3P")
        "Relay"       -> ("Relay", "")   # no digit found
    """
    match = _CLASS_NAME_PATTERN.match(class_name)
    if match:
        name, spec = match.group(1), match.group(2)
        return name.strip(), spec.strip()
    return class_name, ""


def _priority_sort_key(class_name: str):
    """
    Sort key that places element names according to CLASS_PRIORITY_ORDER
    first (in that exact order), and any unlisted element names after,
    sorted alphabetically. Ties within the same element name are broken
    alphabetically by the full class name (so e.g. MCB1P sorts before MCB3P).
    """
    element_name, _ = split_class_name(class_name)
    element_name_upper = element_name.upper()

    priority_lookup = {name.upper(): i for i, name in enumerate(CLASS_PRIORITY_ORDER)}

    if element_name_upper in priority_lookup:
        return (0, priority_lookup[element_name_upper], class_name)
    return (1, element_name_upper, class_name)


def fill_template(
    template_file,
    class_totals: dict,
    panel_name: str = "",
    date: str = "",
    client_name: str = "",
):
    """
    Fills the BOQ template with detection results.

    Args:
        template_file: path to the .xlsx template, or a file-like object
                        (e.g. a Streamlit UploadedFile).
        class_totals: dict of {class_name: total_count}, e.g.
                       {"CB_3Pole": 12, "CB_1Pole": 8, ...}
        panel_name: optional text to write into the "Panel Name" field.
        date: optional text to write into the "Date" field.
        client_name: optional text to write into the "To: Client" field
                     (the "به: شركت" field in the template).

    Returns:
        A BytesIO object containing the filled .xlsx file, ready for download.

    Raises:
        ValueError: if there are more distinct element types than the
                    template has blocks/rows available for.
    """
    wb = openpyxl.load_workbook(template_file)
    ws = wb.active

    # Sort by the priority list defined above (CLASS_PRIORITY_ORDER), not by
    # detection order or count.
    items = sorted(class_totals.items(), key=lambda x: _priority_sort_key(x[0]))

    total_blocks_available = _count_available_blocks(ws)
    blocks_needed = max(1, -(-len(items) // DATA_ROWS_PER_BLOCK))  # ceil division

    if blocks_needed > total_blocks_available:
        raise ValueError(
            f"Not enough blocks in the template: {len(items)} element types need "
            f"{blocks_needed} block(s) of {DATA_ROWS_PER_BLOCK} rows each, but the "
            f"template only has {total_blocks_available} block(s)."
        )

    item_index = 0
    for block_num in range(blocks_needed):
        block_start_row = 1 + block_num * BLOCK_HEIGHT
        header_row = block_start_row + HEADER_OFFSET
        data_start_row = block_start_row + DATA_START_OFFSET

        if panel_name:
            ws.cell(row=block_start_row + 1, column=1).value = f"Panel Name: {panel_name}"
        if date:
            ws.cell(row=block_start_row, column=1).value = f"Date : {date}"
        if client_name:
            ws.cell(row=block_start_row + 1, column=8).value = f"به : {client_name}"

        for row_offset in range(DATA_ROWS_PER_BLOCK):
            if item_index >= len(items):
                break

            class_name, count = items[item_index]
            element_name, spec = split_class_name(class_name)

            row = data_start_row + row_offset
            ws.cell(row=row, column=COL_DESCRIPTION).value = element_name
            ws.cell(row=row, column=COL_RANGE).value = spec
            ws.cell(row=row, column=COL_QTY).value = count

            item_index += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _count_available_blocks(ws):
    """Counts how many repeating blocks exist in the template by scanning
    for 'ITEM' header cells in column A."""
    count = 0
    row = 5
    while row <= ws.max_row:
        if ws.cell(row=row, column=1).value == "ITEM":
            count += 1
        row += BLOCK_HEIGHT
    return count