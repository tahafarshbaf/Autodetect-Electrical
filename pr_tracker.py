"""
PR (Price Request / Proposal) tracking file module.

Appends a new entry to the company's PR tracking Excel file (the
"PR - 1405.xlsx"-style file). The file has two side-by-side tables on
the same sheet; this module only touches the active (right-hand) one:

    Column K = PR number (sequential)
    Column L = تاریخ پیشنهاد (proposal date, Jalali, format "YYYY-MM-DD")
    Column M = نوع پروژه (project type)
    Column N = نام شرکت/پروژه (company/project name)

Rule: always append right after the LAST row that has a PR number in
column K, using (last number + 1) as the new PR number. Any earlier rows
with empty cells are ignored — e.g. if the highest existing PR number is
91, the new entry goes to 92 in the row right after wherever 91 is,
regardless of what's empty above that.
"""

import openpyxl

SHEET_NAME = "Sheet1"

COL_PR_NUMBER = 11   # column K
COL_DATE = 12        # column L
COL_PROJECT_TYPE = 13  # column M
COL_CUSTOMER_NAME = 14  # column N

FIRST_DATA_ROW = 2


def find_last_pr_row(ws, number_col: int = COL_PR_NUMBER, first_row: int = FIRST_DATA_ROW):
    """
    Scans column K from first_row to the sheet's last used row and
    returns (last_row_with_a_number, last_number) for whichever row has
    the LAST (bottom-most) PR number — ignoring any empty rows in
    between.

    If no row has a PR number yet, returns (first_row - 1, 0), so the
    caller's "+1" logic naturally starts at first_row with number 1.
    """
    last_row = first_row - 1
    last_number = 0
    for row in range(first_row, ws.max_row + 1):
        value = ws.cell(row=row, column=number_col).value
        if value not in (None, ""):
            last_row = row
            last_number = value
    return last_row, last_number


def add_pr_entry(file_path: str, project_type: str, customer_name: str, date_str: str):
    """
    Appends one new entry to the PR tracking file at file_path, in the
    row right after the last row that has a PR number.

    Args:
        file_path: path to the PR - 1405.xlsx-style file on disk.
        project_type: text for the "نوع پروژه" column (e.g. "تابلو").
        customer_name: text for the "نام شرکت/پروژه" column.
        date_str: pre-formatted date string for the "تاریخ پیشنهاد"
                   column, e.g. "1405-05-26" (Jalali, dash-separated,
                   matching the existing rows' format).

    Returns:
        (row, pr_number): the row that was written and the PR number
        used, so the caller can show a confirmation message.

    Raises:
        PermissionError: if the file is currently open in Excel (or
                          otherwise locked) and can't be saved.
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb[SHEET_NAME]

    last_row, last_number = find_last_pr_row(ws)
    row = last_row + 1
    pr_number = last_number + 1

    ws.cell(row=row, column=COL_PR_NUMBER).value = pr_number
    ws.cell(row=row, column=COL_DATE).value = date_str
    ws.cell(row=row, column=COL_PROJECT_TYPE).value = project_type
    ws.cell(row=row, column=COL_CUSTOMER_NAME).value = customer_name

    wb.save(file_path)
    return row, pr_number
